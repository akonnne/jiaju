"""生成接口联调检查单 docs/checklist.xlsx。

依据：实施方案 v2.1 第 9 章步骤 7.1（45 接口 × 7 状态列）。
运行：cd api && python scripts/gen_checklist.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.main import app

STATUS_COLS = ["正常", "401", "403", "404", "422", "429", "500"]

METHOD_LABEL = {"get": "GET", "post": "POST", "put": "PUT", "delete": "DELETE", "patch": "PATCH"}


def collect_operations():
    ops = []
    for path, methods in app.openapi()["paths"].items():
        for method, op in methods.items():
            if method.lower() not in METHOD_LABEL:
                continue
            summary = op.get("summary") or op.get("description") or ""
            ops.append((METHOD_LABEL[method.lower()], path, summary))
    # 排序：公开在前，管理在后
    ops.sort(key=lambda x: (x[1].startswith("/api/public"), x[1], x[0]))
    return ops


def main():
    ops = collect_operations()
    wb = Workbook()
    ws = wb.active
    ws.title = "接口联调检查单"

    header = ["#", "方法", "路径", "说明"] + STATUS_COLS + ["备注"]
    ws.append(header)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="7A5C3E")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, (method, path, summary) in enumerate(ops, start=1):
        row = [i, method, path, summary] + [""] * len(STATUS_COLS) + [""]
        ws.append(row)
        ws.cell(row=i + 1, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=i + 1, column=2).alignment = Alignment(horizontal="center")

    widths = [5, 8, 42, 36] + [7] * 7 + [18]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w
    ws.freeze_panes = "A2"

    out = Path(__file__).resolve().parents[2] / "docs" / "checklist.xlsx"
    wb.save(out)
    print(f"✅ 已生成 {out}（{len(ops)} 个接口 × {len(STATUS_COLS)} 状态列）")


if __name__ == "__main__":
    main()
